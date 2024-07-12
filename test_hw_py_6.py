import os
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from conftest import RESOURCES_DIR


def test_for_csv():
    with ZipFile(os.path.join(RESOURCES_DIR, 'zapakovkaslyoz.zip')) as zip_file:
        with zip_file.open('file.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert second_row[0] == 'KFC'
            assert second_row[1] == '784'


def test_for_pdf():
    with ZipFile(os.path.join(RESOURCES_DIR, 'zapakovkaslyoz.zip')) as zip_file:
        with zip_file.open('file.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]
            text = page.extract_text().split('\n')[0]
            assert text == 'Python Testing with pytest'


def test_from_xlsx():
    with ZipFile(os.path.join(RESOURCES_DIR, 'zapakovkaslyoz.zip')) as zip_file:
        with zip_file.open('file.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            second_cell_value = sheet.cell(row=2, column=2).value
            assert second_cell_value == 'Dulce'